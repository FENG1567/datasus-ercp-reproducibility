from __future__ import annotations

"""Extract municipal-level IVS (2000/2010) from the official IPEA Atlas IVS
workbook into a frozen-stage parquet, then prepare a population-weighted
state-level comparison for external validation against Ipeadata state IVS."""

import argparse
import json
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
import pandas as pd

KEY_COLS = [
    "id", "nivel", "ano", "uf", "municipio", "municipio_6digt",
    "ivs", "ivs_infraestrutura_urbana", "ivs_capital_humano", "ivs_renda_e_trabalho",
    "idhm", "idhm_long", "idhm_educ", "idhm_renda",
    "label_cor", "label_sexo", "label_sit_dom",
    "populacao", "renda_per_capita", "i_gini", "t_analf_15m", "t_vulner",
    "t_sem_agua_esgoto", "t_sem_lixo", "t_desocup18m", "espvida",
]
MUNICIPAL_NIVEL = "regiao,uf,rm,municipio"


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--audit", type=Path, required=True)
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.xlsx, read_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col_index = {name: i for i, name in enumerate(header)}
    selected = [c for c in KEY_COLS if c in col_index]
    missing = [c for c in KEY_COLS if c not in col_index]
    rows = []
    municipal_counts = {2000: 0, 2010: 0}
    for raw in ws.iter_rows(min_row=2, values_only=True):
        nivel = raw[col_index["nivel"]]
        ano = raw[col_index["ano"]]
        if nivel != MUNICIPAL_NIVEL:
            continue
        try:
            ano_i = int(ano)
        except (TypeError, ValueError):
            continue
        if ano_i not in (2000, 2010):
            continue
        municipal_counts[ano_i] += 1
        record = {c: raw[col_index[c]] for c in selected}
        rows.append(record)
    wb.close()

    df = pd.DataFrame(rows)
    for col in ["ivs", "ivs_infraestrutura_urbana", "ivs_capital_humano", "ivs_renda_e_trabalho",
                "idhm", "idhm_long", "idhm_educ", "idhm_renda", "populacao", "renda_per_capita",
                "i_gini", "t_analf_15m", "t_vulner", "t_sem_agua_esgoto", "t_sem_lixo",
                "t_desocup18m", "espvida"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col].astype(str).str.replace(",", "."), errors="coerce")

    args.output.parent.mkdir(parents=True, exist_ok=True)
    df.to_parquet(args.output, index=False)
    audit = {
        "schema_version": "1.0",
        "accessed_at": utc_now(),
        "source_workbook": str(args.xlsx),
        "source": "IPEA Atlas IVS official cockpit storage (ivs.ipea.gov.br)",
        "rows_total_municipal": int(len(df)),
        "municipal_by_year": municipal_counts,
        "selected_columns": selected,
        "missing_columns": missing,
        "municipio_code_distinct": int(df["municipio"].nunique()) if "municipio" in df.columns else None,
        "output": str(args.output),
    }
    args.audit.parent.mkdir(parents=True, exist_ok=True)
    args.audit.write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(audit, ensure_ascii=True, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())